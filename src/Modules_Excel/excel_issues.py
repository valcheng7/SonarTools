def patch_time():
    return
import gevent.monkey
gevent.monkey.patch_time = patch_time()
from concurrent.futures import as_completed
from requests_futures.sessions import FuturesSession
import openpyxl
import ast
import pandas as pd 
import seaborn as sns 
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
plt.switch_backend('agg')
import json
import string
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import Reference, Series, BarChart3D
from copy import deepcopy
import time
import datetime
import sys 
import os

username = '180972B'
password = '08032021'

def exportExcel(projectTitle, issues_data, name, params2):

    def resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

    res = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/measures/component?component={projectTitle}&metricKeys=security_rating,reliability_rating,sqale_rating,security_review_rating,security_hotspots,security_hotspots_reviewed,duplicated_lines_density,ncloc_language_distribution,alert_status,code_smells,coverage,bugs,vulnerabilities,reliability_remediation_effort,security_remediation_effort,sqale_index,duplicated_blocks", auth=(username, password))
    sqaleRating, securiyReviewRating, securityRating, reliabilityRating,index = 0, 0, 0, 0,0 
    index = 1
    percentage, num = 0,0 
    duplicatePercentage, linesOfCode = 0,0
    if res.json()["component"]["measures"] != []:
        for i in res.json()["component"]["measures"]:
            if i["metric"] == "sqale_rating":
                sqaleRating = i["value"]
            if i["metric"] == "duplicated_blocks":
                duplicated_blocks = i["value"]
            elif i["metric"] == "reliability_remediation_effort":
                time1 = int(i["value"])
            elif i["metric"] == "security_remediation_effort":
                time2 = int(i["value"])
            elif i["metric"] == "sqale_index":
                time3 = int(i["value"])
            elif i["metric"] == "reliability_rating":
                reliabilityRating = i["value"]
            elif i["metric"] == "coverage":
                coverage = i["value"]
            elif i["metric"] == "security_rating":
                securityRating = i["value"]
            elif i["metric"] == "security_review_rating":
                securiyReviewRating = i["value"]
            elif i['metric'] == 'security_hotspots_reviewed':
                percentage = i['value']
            elif i['metric'] == "security_hotspots":
                num = i['value']
            elif i["metric"] == "duplicated_lines_density":
                duplicatePercentage = i["value"]
            elif i["metric"] == "ncloc":
                linesOfCode = i["value"]
            elif i["metric"] == "code_smells":
                number = i["value"]
            elif i["metric"] == "bugs":
                bugs_number = i["value"]
            elif i["metric"] == "vulnerabilities":
                vul_num = i["value"]
            elif i["metric"] == "ncloc_language_distribution":
                values = i["value"].split(";")
                languages = [{i[:i.index("=")]:i[i.index('=')+1:]} for i in values]
                code_language = []
                for i in languages:
                    code_language.append(list(i.keys())[0])
                total = 0
                for i in languages:
                    total += int(list(i.values())[0])
            elif i["metric"] == "alert_status":
                if i["value"] == "OK":
                    status = "Passed"
                else:
                    status = "Failed"
            proj_id = "a"+str(index)
        projectRatings = {"projectName":res.json()["component"]["key"], "maintainabilityRating":sqaleRating, "securityReviewRating": securiyReviewRating, "reliabilityRating":reliabilityRating, "securityRating":securityRating, "hotspotNumber":num, "hotspotPercentage": percentage, "languages":(languages,total), "duplicatePercentage":duplicatePercentage, "linesOfCodes":linesOfCode, "status":status, "code_smells":number, "coverage":coverage, "bugs":bugs_number, "vulnerabilities":vul_num, "codeLanguages":code_language, "reliabilityEffort":time1, "securityEffort":time2, "maintainabilityEffort":time3, "duplicated_blocks":duplicated_blocks, "proj_id":proj_id, "lines":total}

    dic = {}
    res = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/ce/component?component={projectTitle}", auth=(username, password)) 
    if "current" in res.json():
        projectRatings['lastAnalysisDate'] = res.json()["current"]["submittedAt"][:res.json()["current"]["submittedAt"].index('T')]

    date = projectRatings['lastAnalysisDate']
    indexes = []
    for index, i in enumerate(date):
        if i == '-':
            indexes.append(index)
    year = date[:indexes[0]]
    month = date[indexes[0]+1:indexes[1]]
    day = date[indexes[1]+1:]
    date_formatted = f"{day}/{month}/{year}"
    projectRatings['lastAnalysisDate'] = date_formatted

    def getProjectSeverity(projectTitle):
        project = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/issues/search?facets=severities&componentKeys={projectTitle}&resolved=false", auth=(username, password))
        severity = json.loads(project.text)
        data = severity['facets'][0]['values']
        barGraph = {}
        for i in data:
            if i['val'] == 'MINOR':
                barGraph['MINOR'] = i['count']
            elif i['val'] == 'MAJOR':
                barGraph['MAJOR'] = i['count']
            elif i['val'] == 'INFO':
                barGraph['INFO'] = i['count']
            elif  i['val'] == 'CRITICAL':
                barGraph['CRITICAL'] = i['count']
            else:
                barGraph['BLOCKER'] = i['count']
        return barGraph
        
    def getProjectTypes(projectTitle):
        data1 = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/issues/search?facets=types&componentKeys={projectTitle}&resolved=false", auth=(username, password))
        data2 = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/qualitygates/project_status?projectKey={projectTitle}", auth=(username, password))
        status = json.loads(data2.text)['projectStatus']['status']
        types = json.loads(data1.text)['facets'][0]['values']
        blocks = {}
        if status == 'OK':
            blocks['STATUS'] = "Passed"
        else:
            blocks['STATUS'] = "Failed"
        for i in types:
            if i['val'] == 'CODE_SMELL':
                blocks['CODE_SMELL'] = i['count']
            elif i['val'] == 'BUG':
                blocks['BUG'] = i['count']
            else:
                blocks['VULNERABILITY'] = i['count']
        return blocks

    def getProjectLanguages(projectTitle):
        data3 = requests.get(f"http://sonarqubedev.southeastasia.cloudapp.azure.com/api/issues/search?facets=languages&componentKeys={projectTitle}&resolved=false", auth=(username, password))
        languages = json.loads(data3.text)['facets'][0]['values']
        bar2, total = {}, 0
        for x in languages:
            total += int(x['count'])
        for i in languages:
            percentage = round((i['count'] / total) * 100)
            if i['val'] == 'cs':
                bar2["C#"] = [i['count'], percentage]
            elif i['val'] == 'css':
                bar2["CSS"] = [i['count'], percentage]
            elif i['val'] == 'js':
                bar2["JavaScript"] = [i['count'], percentage]
            elif i['val'] == 'web':
                bar2["HTML"] = [i['count'], percentage]
            elif i['val'] == 'ruby':
                bar2["Ruby"] = [i['count'], percentage]
            elif i['val'] == 'vbnet':
                bar2["VB.NET"] = [i['count'], percentage]
            elif i['val'] == 'xml':
                bar2["XML"] = [i['count'], percentage]
            else:
                bar2[i['val']] = [i['count'], percentage]
        return bar2

    data = projectRatings
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Project Overview')
    wb.create_sheet(index=1, title='Issues')

    ## Removing extra sheet automatically generated ##
    extra_sheet = wb['Sheet']
    wb.remove_sheet(extra_sheet)

    ## Font ##
    font = Font(name='Bahnschrift',size=20,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font1 = Font(name='Bahnschrift',size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font2 = Font(name='Bahnschrift',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font3 = Font(name='Calibri',size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font4 = Font(name='Calibri',size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font5 = Font(name='Bahnschrift',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font6 = Font(name='Bahnschrift',size=12,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font7 = Font(name='Calibri',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    ## Borders ##
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    ## Sheets ##
    first_sheet =  wb['Project Overview']
    second_sheet = wb['Issues']

    """First Sheet"""

    ## Adding Project Overview text ##
    first_sheet['A1'] = 'Project Overview'
    first_sheet.row_dimensions[1].height = 38.4
    first_sheet['A1'].font = font
    first_sheet['A1'].alignment = Alignment(vertical='center')

    ## Adding color to top cells ##
    widths = [19.44, 15.56, 20.67, 13.22, 14.44, 8.78, 17.11, 19.11, 15.11, 12.11, 14.67, 16.11]
    upperCases = string.ascii_uppercase
    for j in range(0,len(widths)):
        first_sheet.column_dimensions[f'{upperCases[j]}'].width = widths[j]
        first_sheet[f'{upperCases[j]}1'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    ## Making the other cells white ##
    for i in range(2, 85):
        first_sheet.row_dimensions[i].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF',fill_type='solid')

    ## Inserting Sonar Tools Icon ##
    ws = wb['Project Overview']
    img = openpyxl.drawing.image.Image(resource_path('src/static/img/brand/cyros_white.png'))
    img.anchor ='J1'
    img.height = 30.72
    img.width = 241.92
    ws.add_image(img)

    ## Adding Name and Date ##
    first_sheet['A3'] = f'Name:  {name}'
    first_sheet['A3'].font = font1
    first_sheet['A3'].alignment = Alignment(vertical='bottom')
    date = str(datetime.datetime.now()).split(' ')[0]
    indexes = []
    for index, i in enumerate(date):
        if i == '-':
            indexes.append(index)

    year = date[:indexes[0]]
    month = date[indexes[0]+1:indexes[1]]
    day = date[indexes[1]+1:]
    date_formatted = f"{day}/{month}/{year}"
    first_sheet['A4'] = f'Date: {date_formatted}'
    first_sheet['A4'].font = font1
    first_sheet['A4'].alignment = Alignment(vertical='bottom')

    ## Project details table ##
    first_sheet.merge_cells('A7:B7')
    first_sheet['A7'] = 'Project Details:'
    first_sheet['A7'].font = font2
    first_sheet['A7'].border=thin_border
    first_sheet['A7'].alignment = Alignment(horizontal='left', vertical='center')
    first_sheet['B7'].border=thin_border
    first_sheet['A7'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')
    first_sheet['B7'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    increment = 0
    titles = ['Project Name', 'Quality Gate Status', 'Last Scanned', 'Hotspot reviewed', 'Coverage', 'Duplication', 'No. of lines of code']
    project_details_list = [projectRatings['projectName'], projectRatings['status'], projectRatings['lastAnalysisDate'], projectRatings['hotspotNumber'], f"{projectRatings['coverage']}%", f"{projectRatings['duplicatePercentage']}%", projectRatings['lines']]

    for i in range(8, 15):
        first_sheet[f'A{i}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        first_sheet[f'A{i}'].border = thin_border
        first_sheet[f'A{i}'].font = font3
        first_sheet[f'A{i}'] = titles[increment]

        first_sheet[f'B{i}'].border = thin_border
        first_sheet[f'B{i}'].font = font4
        first_sheet[f'B{i}'].alignment = Alignment(horizontal='center', vertical='bottom')
        first_sheet[f'B{i}'] = project_details_list[increment]
        
        increment += 1

    ## Security Level Overview ##
    first_sheet['A18'] = 'Severity Level Overview:'
    first_sheet['A18'].font = font5
    first_sheet['A18'].alignment = Alignment(horizontal='left', vertical='center')

    first_sheet['A20'] = 'Severity Level'
    first_sheet['A20'].font = font2
    first_sheet['A20'].border = thin_border
    first_sheet['A20'].alignment = Alignment(horizontal='left', vertical='center')
    first_sheet['A20'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    first_sheet['B20'] = 'No. of issues'
    first_sheet['B20'].font = font2
    first_sheet['B20'].border = thin_border
    first_sheet['B20'].alignment = Alignment(horizontal='left', vertical='center')
    first_sheet['B20'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    increment = 0
    severity_titles = ['Major', 'Minor', 'Critical', 'Blocker', 'Info']
    data = getProjectSeverity(f'{projectTitle}')
    severity_list = [data['MAJOR'], data['MINOR'], data['CRITICAL'], data['BLOCKER'], data['INFO']]

    # rows = [(severity_titles[j], severity_list[j]) for j in range(len(severity_titles))]
    # # data = Reference(wb, min_col=2, min_row=1, max_row=len(severity_titles))
    # # chart = BarChart3D()
    # # chart.title = 'Comparison of severity levels'
    # # chart.add_data(data=data)

    dataset = pd.DataFrame([{"Severity":'MAJOR',"No. of issues":severity_list[0]}, {"Severity":'MINOR',"No. of issues":severity_list[1]},{"Severity":'CRITICAL', "No. of issues":severity_list[2]},{"Severity":'BLOCKER', "No. of issues":severity_list[3]},{"Severity":'INFO', "No. of issues":severity_list[4]}])

    sns.catplot(x="No. of issues", y='Severity', data=dataset, kind='bar')
    plt.savefig('src/static/img/issues_pdf/severity.png', bbox_inches='tight')

    img = openpyxl.drawing.image.Image('src/static/img/issues_pdf/severity.png')
    img.anchor ='A27'
    img.height = 432
    img.width = 617.28
    first_sheet.add_image(img)



    for i in range(21, 26):
        first_sheet[f'A{i}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        first_sheet[f'A{i}'].border = thin_border
        first_sheet[f'A{i}'].font = font3
        first_sheet[f'A{i}'] = severity_titles[increment]
        first_sheet[f'A{i}'].alignment = Alignment(horizontal='center', vertical='center')

        first_sheet[f'B{i}'].border = thin_border
        first_sheet[f'B{i}'].font = font4
        first_sheet[f'B{i}'].alignment = Alignment(horizontal='center', vertical='bottom')
        first_sheet[f'B{i}'] = severity_list[increment]
        
        increment += 1

    ## Types of error table ##
    first_sheet['H18'] = 'Types of error:'
    first_sheet['H18'].font = font5
    first_sheet['H18'].alignment = Alignment(horizontal='left', vertical='center')

    first_sheet['H20'] = 'Type'
    first_sheet['H20'].font = font2
    first_sheet['H20'].border = thin_border
    first_sheet['H20'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['H20'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    first_sheet['I20'] = 'No. of issues'
    first_sheet['I20'].font = font2
    first_sheet['I20'].border = thin_border
    first_sheet['I20'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['I20'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    increment = 0
    type_titles = ['Bugs', 'Vulnerabilities', 'Code Smell']
    data = getProjectTypes(f'{projectTitle}')
    type_list = [data['BUG'], data['VULNERABILITY'], data['CODE_SMELL']]


    for i in range(21, 24):
        first_sheet[f'H{i}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        first_sheet[f'H{i}'].border = thin_border
        first_sheet[f'H{i}'].font = font3
        first_sheet[f'H{i}'] = type_titles[increment]
        first_sheet[f'H{i}'].alignment = Alignment(horizontal='center', vertical='center')

        first_sheet[f'I{i}'].border = thin_border
        first_sheet[f'I{i}'].font = font4
        first_sheet[f'I{i}'].alignment = Alignment(horizontal='center', vertical='bottom')
        first_sheet[f'I{i}'] = type_list[increment]
        
        
        increment += 1

    dataset = pd.DataFrame([{"Type":"Bugs", "No. of issues":type_list[0]}, {"Type":"Vulnerabilities", "No. of issues":type_list[1]}, {"Type":"Code Smell", "No. of issues":type_list[2]}])

    sns.catplot(x='Type', y='No. of issues', data=dataset, kind='bar')

    plt.savefig('src/static/img/issues_pdf/errors.png', bbox_inches='tight')

    img = openpyxl.drawing.image.Image('src/static/img/issues_pdf/errors.png')
    img.anchor ='H27'
    img.height = 432
    img.width = 617.28
    first_sheet.add_image(img)

    ## No.of lines for each language used in this project ##
    first_sheet['A51'] = 'No. of lines for each language used in this project:'
    first_sheet['A51'].font = font5
    first_sheet['A51'].alignment = Alignment(horizontal='left', vertical='center')

    first_sheet['A53'] = 'Language'
    first_sheet['A53'].font = font2
    first_sheet['A53'].border = thin_border
    first_sheet['A53'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['A53'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    first_sheet['B53'] = 'No. of lines'
    first_sheet['B53'].font = font2
    first_sheet['B53'].border = thin_border
    first_sheet['B53'].alignment = Alignment(horizontal='center', vertical='center')
    first_sheet['B53'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    data = getProjectLanguages(f'{projectTitle}')

    count = 54

    for i in data:
        first_sheet[f'A{count}'] = i
        first_sheet[f'A{count}'].border = thin_border
        first_sheet[f'A{count}'].font = font3
        first_sheet[f'A{count}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        first_sheet[f'A{count}'].alignment = Alignment(horizontal='center', vertical='center')

        first_sheet[f'B{count}'] = data[i][0]
        first_sheet[f'B{count}'].border = thin_border
        first_sheet[f'B{count}'].alignment = Alignment(horizontal='center', vertical='bottom')
        first_sheet[f'B{count}'].font = font4
        count += 1

    dataset = pd.DataFrame([{'Language':j,"No. of lines":int(data[j][0])} for j in data])

    sns.catplot(x="No. of lines", y='Language', data=dataset, kind='bar',)
    plt.savefig('src/static/img/issues_pdf/language.png', bbox_inches='tight')

    img = openpyxl.drawing.image.Image('src/static/img/issues_pdf/language.png')
    img.anchor ='A56'
    # img.height = 432
    # img.width = 617.28
    first_sheet.add_image(img)

    # """Second Sheet"""
    # Adding Project Overview text ##
    second_sheet['A1'] = 'Project Issues'
    second_sheet.row_dimensions[1].height = 38.4
    second_sheet['A1'].font = font
    second_sheet['A1'].alignment = Alignment(vertical='center')

    ## Adding color to top cells ##
    widths = [32, 150, 20.67, 13.22, 14.44, 8.78, 17.11, 19.11, 15.11, 12.11, 14.67, 16.11]
    upperCases = string.ascii_uppercase
    for j in range(0,len(widths)):
        second_sheet.column_dimensions[f'{upperCases[j]}'].width = widths[j]
        second_sheet[f'{upperCases[j]}1'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    ## Inserting Sonar Tools Icon ##
    ws = wb['Issues']
    img = openpyxl.drawing.image.Image(resource_path('src/static/img/brand/cyros_white.png'))
    img.anchor ='J1'
    img.height = 30.72
    img.width = 241.92
    ws.add_image(img)

    ## Adding Name and Date ##
    second_sheet['A3'] = f'Name:  {name}'
    second_sheet['A3'].font = font1
    second_sheet['A3'].alignment = Alignment(vertical='bottom')
    date = str(datetime.datetime.now()).split(' ')[0]
    indexes = []
    for index, i in enumerate(date):
        if i == '-':
            indexes.append(index)

    year = date[:indexes[0]]
    month = date[indexes[0]+1:indexes[1]]
    day = date[indexes[1]+1:]
    date_formatted = f"{day}/{month}/{year}"
    second_sheet['A4'] = f'Date: {date_formatted}'
    second_sheet['A4'].font = font1
    second_sheet['A4'].alignment = Alignment(vertical='bottom')

    ## Issues For Table ##
    second_sheet['A6'] = f'Issues For {projectTitle}'
    second_sheet['A6'].font = font6
    second_sheet['A6'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')
    second_sheet['A6'].border = thin_border
    increment_count = 0

    if params2['type'] != []:
        for i in params2['type']:
            second_sheet[f'A{7+increment_count}'] = i
            second_sheet[f'A{7+increment_count}'].font = font1
            second_sheet[f'A{7+increment_count}'].border = thin_border
            second_sheet[f'A{7+increment_count}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
            increment_count += 1
    if params2['date'] != []:
        if len(params2['date']) == 1:
            second_sheet[f'A{7+increment_count}'] = f'From {params2["date"][0]} to {params2["date"][0]}'
            second_sheet[f'A{7+increment_count}'].font = font1
            second_sheet[f'A{7+increment_count}'].border = thin_border
            second_sheet[f'A{7+increment_count}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        else:
            second_sheet[f'A{7+increment_count}'] = f'From {params2["date"][0]} to {params2["date"][1]}'
            second_sheet[f'A{7+increment_count}'].font = font1
            second_sheet[f'A{7+increment_count}'].border = thin_border
            second_sheet[f'A{7+increment_count}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')



    ## Issues ##

    start = 7 + increment_count + 2

    # issues_data = specificProjectSourceCodeIssues(f'{projectTitle}')

    # print(issues_data)

    for i in issues_data:
        second_sheet[f'A{start}'] = f'File Name: {i}'
        second_sheet[f'A{start}'].font = font5
        second_sheet[f'A{start}'].alignment = Alignment(horizontal='left', vertical='center')
        for j in issues_data[i]:
            second_sheet.merge_cells(f'A{start+2}:B{start+2}')
            second_sheet[f'A{start+2}'] = f'Issue'
            second_sheet[f'A{start+2}'].border = thin_border
            second_sheet[f'A{start+2}'].font = font7
            second_sheet[f'A{start+2}'].alignment = Alignment(horizontal='center', vertical='bottom')
            second_sheet[f'A{start+2}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
            second_sheet[f'B{start+2}'].border = thin_border
            second_sheet[f'B{start+2}'].font = font7
            second_sheet[f'B{start+2}'].alignment = Alignment(horizontal='center', vertical='bottom')
            second_sheet[f'B{start+2}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
            li = ['Message', 'Type', 'Severity', 'Status', 'Effort', 'Date', 'Line']
            keys = ['message', 'type', 'severity', 'status', 'effort', 'creationDate', 'startLine']
            start += 2
            for k in range(1,8):
                second_sheet[f'A{start+k}'] = f'{li[k-1]}'
                second_sheet[f'A{start+k}'].font = font3
                second_sheet[f'A{start+k}'].border = thin_border
                second_sheet[f'A{start+k}'].alignment = Alignment(horizontal='center', vertical='bottom')
                second_sheet[f'A{start+k}'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')

                if k != 1:
                    second_sheet[f'B{start+k}'] = f'{j[keys[k-1]]}'
                    second_sheet[f'B{start+k}'].font = font4
                    second_sheet[f'B{start+k}'].border = thin_border
                    second_sheet[f'B{start+k}'].alignment = Alignment(horizontal='center', vertical='bottom')
                else:
                    second_sheet[f'B{start+k}'] = f'{j[keys[k-1]]}'
                    second_sheet[f'B{start+k}'].font = font4
                    second_sheet[f'B{start+k}'].border = thin_border
                    second_sheet[f'B{start+k}'].alignment = Alignment(horizontal='center', vertical='bottom')
            second_sheet[f'A{start+9}'] = f'Why is this an issue?'
            second_sheet[f'A{start+9}'].font = font3
            # second_sheet[f'A{start+2}'].border = thin_border
            # second_sheet[f'A{start+2}'].alignment = Alignment(horizontal='center', vertical='bottom')
            start += 3
            ws = wb['Issues']
            print(j['rule'])
            imageName = j['rule'].replace(':', '')
            if '/' in imageName:
                imageName = imageName.replace('/','')
            img = openpyxl.drawing.image.Image(f"src/rules/{imageName}.png")
            img.anchor = f'A{start+9}'
            ws.add_image(img)
            start += 65
        start += 2          

    for i in range(2, start+1):
        second_sheet.row_dimensions[i].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF',fill_type='solid')

    wb.save('src/static/downloads/csv/issues.xlsx')

# exportExcel('webgoat_sonarqube3')