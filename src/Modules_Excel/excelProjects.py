def patch_time():
    return
import gevent.monkey
gevent.monkey.patch_time = patch_time()
from concurrent.futures import as_completed
from requests_futures.sessions import FuturesSession
import openpyxl
import ast
import pandas as pd 
import seaborn as sns 
import matplotlib.pyplot as plt
import json
import string
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.chart import Reference, Series, BarChart3D
from copy import deepcopy
import time
import datetime

import sys
import os

username = '180972B'
password = '08032021'

def exportExcelProjects(name, projectRatings):

    def resource_path(relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='Project Metrics')

    ## Removing extra sheet automatically generated ##
    extra_sheet = wb['Sheet']
    wb.remove_sheet(extra_sheet)

    ## Font ##
    font = Font(name='Bahnschrift',size=20,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font1 = Font(name='Bahnschrift',size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font2 = Font(name='Bahnschrift',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font3 = Font(name='Calibri',size=11,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font4 = Font(name='Calibri',size=11,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font5 = Font(name='Bahnschrift',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    font6 = Font(name='Bahnschrift',size=12,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='00FFFFFF')

    font7 = Font(name='Calibri',size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')

    ## Borders ##
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    ## Sheets ##
    first_sheet =  wb['Project Metrics']

    """First Sheet"""

    ## Adding Project Metrics text ##
    first_sheet['A1'] = 'Project Metrics'
    first_sheet.row_dimensions[1].height = 38.4
    first_sheet['A1'].font = font
    first_sheet['A1'].alignment = Alignment(vertical='center')

    ## Adding color to top cells ##
    widths = [25, 20, 20.67, 20, 20, 20, 20, 20,20, 20, 20,20]
    upperCases = string.ascii_uppercase
    for j in range(0,len(widths)):
        first_sheet.column_dimensions[f'{upperCases[j]}'].width = widths[j]
        first_sheet[f'{upperCases[j]}1'].fill = PatternFill(start_color='5E72E4', end_color='5E72E4',fill_type='solid')

    ## Making the other cells white ##
    for i in range(2, 85):
        first_sheet.row_dimensions[i].fill = PatternFill(start_color='00FFFFFF', end_color='00FFFFFF',fill_type='solid')

    ## Inserting Sonar Tools Icon ##
    ws = wb['Project Metrics']
    img = openpyxl.drawing.image.Image(resource_path('src/static/img/brand/cyros_white.png'))
    img.anchor ='J1'
    img.height = 30.72
    img.width = 241.92
    ws.add_image(img)

    ## Adding Name and Date ##
    first_sheet['A3'] = f'Name:  {name}'
    first_sheet['A3'].font = font1
    first_sheet['A3'].alignment = Alignment(vertical='bottom')
    date = str(datetime.datetime.now()).split(' ')[0]
    indexes = []
    for index, i in enumerate(date):
        if i == '-':
            indexes.append(index)

    year = date[:indexes[0]]
    month = date[indexes[0]+1:indexes[1]]
    day = date[indexes[1]+1:]
    date_formatted = f"{day}/{month}/{year}"
    first_sheet['A4'] = f'Date: {date_formatted}'
    first_sheet['A4'].font = font1
    first_sheet['A4'].alignment = Alignment(vertical='bottom')

    first_sheet.merge_cells('A7:B7')
    first_sheet['A7'] = 'Project Metrics:'
    first_sheet['A7'].font = font5
    first_sheet['A7'].alignment = Alignment(horizontal='left', vertical='center')

    titles = ['Project', 'Quality Gate Status', 'Bugs', 'Vulnerabilities', 'Hotspots Reviewed', 'Code Smells', 'Coverage', 'Duplications', 'No.of lines']
    
    caps = string.ascii_uppercase
    increment = 0

    for i in titles:
        first_sheet[f'{caps[increment]}8'].fill = PatternFill(start_color='B4BDF2', end_color='B4BDF2',fill_type='solid')
        first_sheet[f'{caps[increment]}8'].border = thin_border
        first_sheet[f'{caps[increment]}8'].font = font3
        first_sheet[f'{caps[increment]}8'] = i
        first_sheet[f'{caps[increment]}8'].alignment = Alignment(horizontal='center', vertical='bottom')

        increment += 1

    ## starting index ##
    i = 9

    for j in projectRatings:
        li = [j["projectName"], j["status"], j["bugs"], j["vulnerabilities"], f'{j["hotspotPercentage"]}%', j["code_smells"],f'{j["coverage"]}%', f'{j["duplicatePercentage"]}%', f'{j["languages"][1]} lines'] 
        for k in range(0, 9):      
            first_sheet[f'{caps[k]}{i}'].border = thin_border
            first_sheet[f'{caps[k]}{i}'].font = font4
            first_sheet[f'{caps[k]}{i}'].alignment = Alignment(horizontal='center', vertical='bottom')
            first_sheet[f'{caps[k]}{i}'] = li[k]
        i += 1

    wb.save('src/static/downloads/csv/projects.xlsx')


    